import streamlit as st
import openpyxl as op
import os
from datetime import datetime
from openpyxl.styles import Border,Side

def findoctant(x,y,z): #function for finding octact in which x,y,z lies
	if x>=0:
		if y>=0:
			return 1 if z>=0 else -1
		else:
			return 4 if z>=0 else -4
	else:
		if y>=0:
			return 2 if z>=0 else -2
		else:
			return 3 if z>=0 else -3

def get_value(element): # a function to sort no of octants which will be used in identifying rank
	return element["value"]

def octantAnalysis(m,mod=5000):
	octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
	
	fill_cell= op.styles.PatternFill(patternType='solid', fgColor='FFFF00') # to fill yellow colour
	
	thin = Side(border_style="thin", color="000000") # for border part
	
	try: # opening the input excel sheet
		os.chdir(r'C:\Users\Aman Jha\Documents\GitHub\octantAnalyis\input') # making the parent directory as input directory
		inp=op.load_workbook(m)
		ipsheet=inp.active
	except Exception as e: # handling exception and exitting the program
		print(f"Can't open input file {m}")
		print(e)
		return
	
	try: # opening the input excel sheet
		os.chdir(r'C:\Users\Aman Jha\Documents\GitHub\octantAnalyis\output') # making the parent directory as input directory
		output=op.Workbook()
		opsheet=output.active
	except Exception as e: # handling exception and exitting the program
		print(f"Can't open output file for input file {m}")
		print(e)
		return
	
	# file name
	now=datetime.now()
	if type(m)==str:
		filename=m[:-5]
	else:
		filename=m.name[:-5]
	filename+="_"+str(mod)+"_"+str(now.year)+"_"+str(now.month)+"_"+str(now.day)+"_"+str(now.hour)+"_"+str(now.minute)+"_"+str(now.second)+".xlsx"
	
	# common part of every tut
	try: # copying the column headers from input file
		opsheet['A2'].value=ipsheet['A1'].value 
		opsheet['B2'].value=ipsheet['B1'].value
		opsheet['C2'].value=ipsheet['C1'].value
		opsheet['D2'].value=ipsheet['D1'].value
	except Exception as e: # handling exception and exitting the program
		print(e)
		output.save(filename)
		return

	size=ipsheet.max_row; 
	sumu=0; sumv=0; sumw=0 # variables to sum the values

	for i in range(2,size+1): # copying the value of first four columns from input file and summing the values of u, v and w for calculating average
		opsheet.cell(row=i+1,column=1).value=float(ipsheet.cell(row=i,column=1).value) 
		opsheet.cell(row=i+1,column=2).value=round(float(ipsheet.cell(row=i,column=2).value),3)   
		opsheet.cell(row=i+1,column=3).value=round(float(ipsheet.cell(row=i,column=3).value),3)    
		opsheet.cell(row=i+1,column=4).value=round(float(ipsheet.cell(row=i,column=4).value),3)    

		sumu+=float(ipsheet.cell(row=i,column=2).value)
		sumv+=float(ipsheet.cell(row=i,column=3).value)
		sumw+=float(ipsheet.cell(row=i,column=4).value)
	
	try: # calculating averages
		avgu=round(sumu/(size-1),3)
		avgv=round(sumv/(size-1),3)
		avgw=round(sumw/(size-1),3)
	except Exception as e: # handling exception and exitting the program
		print(e)
		output.save(filename)
		return
	
	# creating columns for average values and putting values in that
	opsheet['E2']='U Avg'; opsheet['E3']=avgu
	opsheet['F2']='V Avg'; opsheet['F3']=avgv
	opsheet['G2']='W Avg'; opsheet['G3']=avgw

	# creating columns for updated values of u,v,w and column of octact
	opsheet['H2']="U'=U- U Avg"
	opsheet['I2']="V'=V- V Avg"
	opsheet['J2']="W'=W- W Avg"
	opsheet['K2']="Octact"

	octcnt={1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0} # to store the overall octant count
	
	# to store the overall transition of octant count
	transcnt={1:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},2:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},3:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},4:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-1:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-2:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-3:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-4:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}}
	
	for i in range(2,size+1): # storing values of updated u,v,w column
		opsheet.cell(row=i+1,column=8).value=float(ipsheet.cell(row=i,column=2).value)-avgu   
		opsheet.cell(row=i+1,column=9).value=float(ipsheet.cell(row=i,column=3).value)-avgv
		opsheet.cell(row=i+1,column=10).value=float(ipsheet.cell(row=i,column=4).value)-avgw

		try:
			y=findoctant(opsheet.cell(row=i+1,column=8).value,opsheet.cell(row=i+1,column=9).value,opsheet.cell(row=i+1,column=10).value) # calling findoctant function
		except Exception as e: # handling exception and exitting the program
			print(e)
			output.save(filename)
			return
		opsheet.cell(row=i+1,column=11).value=y # octact column
		octcnt[y]+=1 # increasing the count of that octant
		if(i==2): 
			prev=y 
			continue # for transition count we need previous value as well, therefore moving on to the next transition count
		transcnt[prev][y]+=1 # maintaining the overall transition count
		prev=y # storing the previous value of octant
	
	# part of 5th tut
	
	opsheet['N1'].value="Overall Octant Count"
	opsheet['N3'].value="Octant ID"
	opsheet['N4'].value="Overall Count" # storing the overall octact count 
	opsheet['M4'].value="Mod " + str(mod)

	ordl=[] # creating an ordered list
	for key,value in octcnt.items():
		ordl.append({"value":value,"key":key})
		
	ordl.sort(key=get_value,reverse=True)

	ordcnt={} # ordered octant count
	rank1id=[] # id of rank1
	rank1id.append([])
	i=0; prev=-1
	for item in ordl:
		if(prev!=item["value"]): i+=1
		if(i==1): rank1id[-1].append(item["key"])
		prev=item["value"]
		ordcnt.update({item["key"]:i})
	
	k=1
	for i in range(15,23,2):
		opsheet.cell(row=3,column=i).value="+"+str(k)
		opsheet.cell(row=4,column=i).value=octcnt[k]

		opsheet.cell(row=3,column=i+1).value=str(-k)
		opsheet.cell(row=4,column=i+1).value=octcnt[-k]
		
		k+=1
	
	k=1
	for i in range(1,9,2):
		opsheet.cell(row=3,column=22+i).value="Rank Octant +"+str(k)
		opsheet.cell(row=4,column=22+i).value=ordcnt[k]

		opsheet.cell(row=3,column=22+i+1).value="Rank Octant -"+str(k)
		opsheet.cell(row=4,column=22+i+1).value=ordcnt[-k]
		
		if ordcnt[k]==1: opsheet.cell(row=4,column=22+i).fill=fill_cell
		if ordcnt[-k]==1: opsheet.cell(row=4,column=22+i+1).fill=fill_cell

		k+=1
	
	opsheet.cell(row=3,column=31).value="Rank 1 Octant ID"
	opsheet.cell(row=4,column=31).value=str(rank1id[-1])[1:-1] # there might be more than one ID having the same rank, and as it is in list format so when we convert to string, we have to remove the brackets
	opsheet.cell(row=3,column=32).value="Rank 1 Octant Name"
	opsheet.cell(row=4,column=32).value=str([octant_name_id_mapping[str(k)] for k in rank1id[-1]])[1:-1]
	rank1id.clear()
	r=5
	
	dic={1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0} # to maintain the count of octants which has rank 1 for the below table of the 1st part of this tut
	# for counting octact in different region
	for i in range(1,size+1,mod): # for loop by increasing the value by mod in 1 iteration
		if(i+mod>=size): # last case when the elements are less than or equal to mod
			last=size
		else: # setting last value of that range
			last=i+mod
		
		opsheet.cell(row=r,column=14).value=str(i-1)+"-"+str(last-2) # filling range in column M

		tempcnt={1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}
		for j in range(i+1,last+1):
			try:
				tempcnt[opsheet.cell(row=j+1,column=11).value]+=1
			except Exception as e: # handling exception and exitting the program after saving it
				print(e)
				output.save(filename)
				return

		k=1
		for z in range(15,23,2):
			opsheet.cell(row=r,column=z).value=tempcnt[k]
			opsheet.cell(row=r,column=z+1).value=tempcnt[-k]
			k+=1

		ordl=[] # creating an ordered list
		for key,value in tempcnt.items():
			ordl.append({"value":value,"key":key})
			
		ordl.sort(key=get_value,reverse=True)

		ordcnt={} # ordered octant count
		rank1id.append([])
		a=0; prev=-1
		for item in ordl:
			if(prev!=item["value"]): a+=1
			if(a==1): 
				rank1id[-1].append(item["key"])
				dic[item["key"]]+=1
			prev=item["value"]
			ordcnt.update({item["key"]:a})

		k=1
		for a in range(1,9,2):
			opsheet.cell(row=r,column=22+a).value=ordcnt[k]
			opsheet.cell(row=r,column=22+a+1).value=ordcnt[-k]
			
			if ordcnt[k]==1: opsheet.cell(row=r,column=22+a).fill=fill_cell
			if ordcnt[-k]==1: opsheet.cell(row=r,column=22+a+1).fill=fill_cell
			k+=1
		
		opsheet.cell(row=r,column=31).value=str(rank1id[-1])[1:-1]
		opsheet.cell(row=r,column=32).value=str([octant_name_id_mapping[str(k)] for k in rank1id[-1]])[1:-1]

		r+=1
	
	r+=1
	
	opsheet.cell(row=r,column=29).value="Octant ID"
	opsheet.cell(row=r,column=30).value="Octant Name"
	opsheet.cell(row=r,column=31).value="Count of Rank 1 Mod Values"
	r+=1
	
	k=1
	for i in range(0,8,2):
		opsheet.cell(row=r,column=29).value="+"+str(k)
		opsheet.cell(row=r,column=30).value=octant_name_id_mapping[str(k)]
		opsheet.cell(row=r,column=31).value=dic[k]

		r+=1
		opsheet.cell(row=r,column=29).value=str(-k)
		opsheet.cell(row=r,column=30).value=octant_name_id_mapping[str(-k)]
		opsheet.cell(row=r,column=31).value=dic[-k]
		r+=1; k+=1
	
	temp=0
	if (size-1)%mod: temp=1
	for row_no in range(3,4+(size-1)//mod+temp+1):
		for col_no in range(14,33):
			opsheet.cell(row=row_no,column=col_no).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
		
	for row_no in range(4+(size-1)//mod+temp+2,4+(size-1)//mod+temp+2+9):
		for col_no in range(29,32):
			opsheet.cell(row=row_no,column=col_no).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
	
	# part of 2nd tut
	# beginning of transition count
	r=1
	opsheet.cell(row=r,column=35).value="Overall Transition Count"

	r+=1
	opsheet.cell(row=r,column=36).value="To"

	r+=1
	opsheet.cell(row=r,column=35).value="Octant #"
	k=1
	for z in range(36,44,2):
		opsheet.cell(row=r,column=z).value="+"+str(k)
		opsheet.cell(row=r,column=z+1).value=str(-k)
		k+=1
	
	r+=1
	opsheet.cell(row=r,column=34).value="From"
	# overall transition count
	k=1
	for j in range(r,r+8):
		l=1
		for z in range(36,44,2):
			opsheet.cell(row=j,column=z).value=transcnt[k][l]
			opsheet.cell(row=j,column=z+1).value=transcnt[k][-l]
			l+=1

		if(k>0):
			opsheet.cell(row=j,column=35).value="+"+str(k) 
			k*=-1
		else:
			opsheet.cell(row=j,column=35).value=str(k) 
			k=k*-1+1

	r=r+11 # increasing by 8 bcoz of the 8 octants and 2 empty rows 
	
	# transition count for specific region
	for i in range(1,size+1,mod): # for loop by increasing the value by mod in 1 iteration
		if(i+mod>=size): # last case when the elements are less than or equal to mod
			last=size
		else: # setting last value of that range
			last=i+mod

		opsheet.cell(row=r,column=35).value="Mod Transition Count"

		r+=1
		opsheet.cell(row=r,column=35).value=str(i-1)+"-"+str(last-2) # filling range in column M
		opsheet.cell(row=r,column=36).value="To"

		r+=1
		opsheet.cell(row=r,column=35).value="Ocatnt #"
		
		k=1
		for z in range(36,44,2):
			opsheet.cell(row=r,column=z).value="+"+str(k)
			opsheet.cell(row=r,column=z+1).value=str(-k)
			k+=1
		
		r+=1
		opsheet.cell(row=r,column=34).value="From"

		temptranscnt={1:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},2:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},3:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},4:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-1:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-2:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-3:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0},-4:{1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}}

		for j in range(i+2,last+2):
			if j==i+2: # for transition count we need previous value as well, therefore moving on to the next transition count
				prev=opsheet.cell(row=j,column=11).value
				continue
			y=opsheet.cell(row=j,column=11).value
			temptranscnt[prev][y]+=1 # maintaining the transition count
			prev=y
			if j==last+1:
				if j==size+1:
					continue
				temptranscnt[y][opsheet.cell(row=j+1,column=11).value]+=1
			
		k=1
		for j in range(r,r+8):
			
			l=1
			for z in range(36,44,2):
				opsheet.cell(row=j,column=z).value=temptranscnt[k][l]
				opsheet.cell(row=j,column=z+1).value=temptranscnt[k][-l]
				l+=1

			# k=-1*k
			if(k>0):
				opsheet.cell(row=j,column=35).value="+"+str(k) 
				k*=-1
			else:
				opsheet.cell(row=j,column=35).value=str(k) 
				k=k*-1+1
		
		r+=11
	r=3
	for k in range((size-1)//mod+temp+1):
		for row_no in range(r,r+9):
			for col_no in range(35,44):
				opsheet.cell(row=row_no,column=col_no).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
		r+=14
	
	# tut 4 part
	opsheet['AS3'].value="Octant"
	opsheet['AS1'].value="Longest Subsequence Length"
	opsheet['AT3'].value="Longest Subsequence Length"
	opsheet['AU3'].value="Count"

	lsl={1:{"maxi":0,"count":0,"li":[]},2:{"maxi":0,"count":0,"li":[]},3:{"maxi":0,"count":0,"li":[]},4:{"maxi":0,"count":0,"li":[]},-1:{"maxi":0,"count":0,"li":[]},-2:{"maxi":0,"count":0,"li":[]},-3:{"maxi":0,"count":0,"li":[]},-4:{"maxi":0,"count":0,"li":[]}} # adding a 3rd element in every sub-dictionary of lsl dictionary to keep the start and end time of all the instance when they have longest subsequence count

	lenth=1; prev=0; tprev=0 

	for i in range(3,size+2):
		y=opsheet.cell(row=i,column=11).value
		if i==3: # 1st data
			prev=y
			continue

		if prev==y:
			lenth+=1
		else:
			if(lenth==lsl[prev]["maxi"]):
				lsl[prev]["count"]+=1
				try:
					lsl[prev]["li"].append([tprev,opsheet.cell(row=i-1,column=1).value])
				except Exception as e: # handling exception and exitting the program
					print(e)
					output.save(filename)
					return
			elif(lenth>lsl[prev]["maxi"]):
				lsl[prev]["maxi"]=lenth
				lsl[prev]["count"]=1

				try:
					lsl[prev]["li"].clear()
					lsl[prev]["li"].append([tprev,opsheet.cell(row=i-1,column=1).value])
				except Exception as e: # handling exception and exitting the program
					print(e)
					output.save(filename)
					return

			lenth=1
			prev=y
			tprev=opsheet.cell(row=i,column=1).value

	if(lenth==lsl[prev]["maxi"]):
		lsl[prev]["count"]+=1
		try:
			lsl[prev]["li"].append([tprev,opsheet.cell(row=size+1,column=1).value])
		except Exception as e: # handling exception and exitting the program
			print(e)
			output.save(filename)
			return

	elif(lenth>lsl[prev]["maxi"]):
		lsl[prev]["maxi"]=lenth
		lsl[prev]["count"]=1
		try:
			lsl[prev]["li"].clear()
			lsl[prev]["li"].append([tprev,opsheet.cell(row=size+1,column=1).value])
		except Exception as e: # handling exception and exitting the program
			print(e)
			output.save(filename)
			return
	
	k=1; tot=0
	for i in range(4,12,2):
		opsheet.cell(row=i,column=45).value="+"+str(k)
		opsheet.cell(row=i,column=46).value=lsl[k]["maxi"]
		opsheet.cell(row=i,column=47).value=lsl[k]["count"]
		tot+=lsl[k]["count"]

		opsheet.cell(row=i+1,column=45).value=str(-k)
		opsheet.cell(row=i+1,column=46).value=lsl[-k]["maxi"]
		opsheet.cell(row=i+1,column=47).value=lsl[-k]["count"]
		tot+=lsl[-k]["count"]

		k+=1
		
	for row_no in range(3,12):
		for col_no in range(45,48):
			opsheet.cell(row=row_no,column=col_no).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
			
	opsheet['AW1'].value="Longest Subsequence Length with Range"
	opsheet['AW3'].value="Octant"
	opsheet['AX3'].value="Longest Subsequence Length"
	opsheet['AY3'].value="Count"

	k=1; r=4
	while r<20+tot: # 20 + tot= 2(lines above)+ 1 (for header of columns) + 2*8 (2 rows for each octact's header and data) + tot (number of time range) + 1 (becuase range(a,b) excludes the last number)
		for j in range(1,3):
			if (j==1): # for positive octact value
				opsheet.cell(row=r,column=49).value="+"+str(k)
			else:
				k=-k
				opsheet.cell(row=r,column=49).value=str(k)
			opsheet.cell(row=r,column=50).value=lsl[k]["maxi"]
			opsheet.cell(row=r,column=51).value=lsl[k]["count"]

			r+=1
			opsheet.cell(row=r,column=49).value="Time"
			opsheet.cell(row=r,column=50).value="From"
			opsheet.cell(row=r,column=51).value="To"

			r+=1
			for i in range(0,lsl[k]["count"]): 
				opsheet.cell(row=r+i,column=50).value=lsl[k]["li"][i][0]
				opsheet.cell(row=r+i,column=51).value=lsl[k]["li"][i][1]
			
			r+=lsl[k]["count"]
			
			if(k<0): k=-k

		k+=1
	
	for row_no in range(3,20+tot):
		for col_no in range(49,52):
			opsheet.cell(row=row_no,column=col_no).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
			
	output.save(filename)
	return
		
st.set_page_config(page_icon=":open_hands:",layout="wide") # basic page configuration

st.title("Hello Sir!!! :wave:")
st.header("Welcome to bulk conversion file page. :v:")
st.write("If you want to convert multiple input files into output files, then you are on the correct page. :heavy_check_mark:")

st.write("Otherwise move to other page. :heavy_multiplication_x:")

st.header("We can take input files by two method:")
st.subheader("1. Uploading files method:")
input_files=st.file_uploader("Upload the input files",type=["xlsx"], accept_multiple_files=True)

st.write("##")

st.subheader("2. Taking path of the files method:")
input_path=st.text_input("Upload the path of the input files",placeholder=r"Please enter the path, for example:- 'C:\Users\Aman Jha\Documents\GitHub\octantAnalyis\\input' (without inverted commas)")

st.header("Taking mod value.")

# dividing page in 2 columns in the ratio 2:1
left_col,right_col=st.columns((2,1))

mod=left_col.text_input("Enter the mod value:",placeholder="Please enter an integer mod value, for example:- 5000")

try:
	mod=int(mod)
except Exception as e:
	st.write("Please enter an integer value for mod")
	print(e)

st.write("##")
work=st.button("Compute",help="Click here for computing the files",type="primary")

st.write("##")
# download=st.button("Download",help="Click here for downloading the file",type="primary")
# location=st.text_input("If you want to download the files in a specific path then please state, otherwise it will get saved in output_files folder in project 2:",placeholder=r"Please enter the path, for example:- 'C:\Users\Aman Jha\Documents\GitHub\2001EE22_2022\tut07\input' (without inverted commas)",value="")

if work:
	try:
		os.mkdir(r'C:\Users\Aman Jha\Documents\GitHub\octantAnalysis\output')
	except:
		pass
	
	for files in input_files:
		octantAnalysis(files,mod)
	
	try:
		x=os.listdir(input_path)
		for files in x:
			if files[-5:]==".xlsx":
				octantAnalysis(files,mod)
	except:
		if input_path != "":
			st.write("Please provide a correct path!!!")
						
	st.success("Computing successful!!!")